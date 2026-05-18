
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def localizar_linha_cabecalho(df):
    for i in range(min(20, len(df))):
        linha = " ".join([str(x).upper() for x in df.iloc[i].tolist()])

        if (
            "PACIENTE" in linha
            and "PROCESSO" in linha
            and "SENHA" in linha
        ):
            return i

    return 0



# TRECHO PRINCIPAL DE NORMALIZAÇÃO DAS COLUNAS


def normalizar_colunas(colunas):
    return [
        "Nome do Paciente",
        "Fila",
        "Especialidade",
        "Classificação",
        "Prestador",
        "Cód. Atendimento",
        "Cód. Triagem",
        "Cód. Paciente",
        "Senha",
        "Data e hora retirada Senha",
        "Data e Hora Processo Inicial",
        "Data e hora processo Final",
        "Data e hora Remoção",
        "Diferença entre Processos"
    ]


def ajustar_dataframe(df):
    quantidade = len(df.columns)

    nomes = normalizar_colunas(df.columns)

    while len(nomes) < quantidade:
        nomes.append(f"Extra {len(nomes)+1}")

    df.columns = nomes[:quantidade]

    # força existir as colunas finais corretas
    if "Data e hora Remoção" not in df.columns:
        df["Data e hora Remoção"] = ""

    if "Diferença entre Processos" not in df.columns:
        df["Diferença entre Processos"] = ""

    return df


def processar():
    arquivo = filedialog.askopenfilename(
        title="Selecione o relatório",
        filetypes=[
            ("Excel", "*.xlsx *.xls"),
            ("CSV", "*.csv")
        ]
    )

    if not arquivo:
        return

    try:
        caminho = Path(arquivo)

        if caminho.suffix.lower() == ".csv":
            bruto = pd.read_csv(
                arquivo,
                header=None,
                sep=None,
                engine="python"
            )
        else:
            bruto = pd.read_excel(
                arquivo,
                header=None
            )

        linha_header = localizar_linha_cabecalho(bruto)

        if caminho.suffix.lower() == ".csv":
            df = pd.read_csv(
                arquivo,
                header=linha_header,
                sep=None,
                engine="python"
            )
        else:
            df = pd.read_excel(
                arquivo,
                header=linha_header
            )

        df = df.dropna(axis=1, how="all")

        df = ajustar_dataframe(df)

        coluna_data = "Data e Hora Processo Inicial"

        if coluna_data not in df.columns:
            raise Exception(
                "Não encontrei a coluna de Processo Inicial."
            )

        df[coluna_data] = pd.to_datetime(
            df[coluna_data],
            dayfirst=True,
            errors="coerce"
        )

        df = df.dropna(subset=[coluna_data])

        df = df.sort_values(
            by=coluna_data,
            ascending=True
        )

        saida = caminho.with_name(
            caminho.stem + "_ORDENADO.xlsx"
        )

        df.to_excel(
            saida,
            index=False,
            startrow=2
        )

        wb = load_workbook(saida)
        ws = wb.active

        hospital = "Nome Sistema"

        ultima_coluna = ws.max_column
        ultima_letra = get_column_letter(ultima_coluna)

        ws.merge_cells(f"A1:{ultima_letra}1")

        titulo = ws["A1"] = ""
        titulo.value = hospital

        titulo.font = Font(
            bold=True,
            size=16
        )

        titulo.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        titulo.fill = PatternFill(
            start_color="D9D9D9",
            end_color="D9D9D9",
            fill_type="solid"
        )

        ws.row_dimensions[1].height = 28

        cabecalho_fill = PatternFill(
            start_color="EDEDED",
            end_color="EDEDED",
            fill_type="solid"
        )

        cabecalho_font = Font(
            bold=True
        )

        borda = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[3]:
            cell.fill = cabecalho_fill
            cell.font = cabecalho_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = borda

        for coluna in range(1, ws.max_column + 1):
            letra = get_column_letter(coluna)

            tamanho_max = 0

            for linha in range(1, ws.max_row + 1):
                valor = ws.cell(linha, coluna).value

                if valor:
                    tamanho_max = max(
                        tamanho_max,
                        len(str(valor))
                    )

            ws.column_dimensions[letra].width = min(tamanho_max + 4, 40)

        ws.freeze_panes = "A4"

        wb.save(saida)

        messagebox.showinfo(
            "Concluído",
            f"Arquivo ordenado com sucesso!\n\n{saida}"
        )

    except Exception as e:
        messagebox.showerror(
            "Erro",
            str(e)
        )


janela = tk.Tk()
janela.title("Organizador de excel")
janela.geometry("550x260")
janela.resizable(False, False)

titulo = tk.Label(
    janela,
    text="Organizador Automático de Relatórios",
    font=("Arial", 16, "bold")
)
titulo.pack(pady=20)

descricao = tk.Label(
    janela,
    text=(
        "Importe o relatório hospitalar.\n"
        "O sistema detecta automaticamente o cabeçalho,\n"
        "padroniza colunas e ordena cronologicamente."
    ),
    font=("Arial", 10)
)
descricao.pack(pady=10)

botao = tk.Button(
    janela,
    text="Selecionar Arquivo",
    width=25,
    height=2,
    command=processar,
    font=("Arial", 11, "bold")
)

botao.pack(pady=20)

janela.mainloop()
